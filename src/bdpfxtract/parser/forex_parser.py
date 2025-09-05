from __future__ import annotations

import datetime as dt
import re
from decimal import Decimal
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from openpyxl.worksheet.worksheet import Worksheet

from bdpfxtract.logging import configure_logging
from bdpfxtract.model import ForexRecord

logger = configure_logging()


class HeaderParser:
    """Parses the worksheet header to map column indexes to currency codes.

    Expected layout (as observed in Banco de Portugal downloads):
    - Row 1: Series IDs (ignored)
    - Row 2: Series titles like "Dólar dos EUA (USD) - diário"
    - Row 3: "Métrica" / "Valor" (ignored)
    - Row 4: "Unidade de Medida" / "Divisa" (ignored)
    - Row 5+: Data rows: Col A is date, Col B.. are numeric values per currency.
    """

    CODE_RE = re.compile(r"\(([A-Z]{3})\)")

    def build_currency_map(self, ws: Worksheet) -> dict[int, str]:
        mapping: dict[int, str] = {}
        header_row = 2
        # Start from column 2 (B) because column 1 (A) is the date column.
        for col_idx in range(2, ws.max_column + 1):
            cell = ws.cell(row=header_row, column=col_idx)
            value = cell.value
            if not value:
                continue
            if not isinstance(value, str):
                continue
            m = self.CODE_RE.search(value)
            if not m:
                # No code found; skip this column to be safe.
                logger.debug(
                    "Skipping column %s without ISO code in header: %r", col_idx, value
                )
                continue
            code = m.group(1)
            mapping[col_idx] = code
        if not mapping:
            raise ValueError("No currency headers found in row 2.")
        logger.info("Detected %d currency columns.", len(mapping))
        return mapping


class RowParser:
    """Parses data rows into ForexRecord entities."""

    def __init__(self, ws_epoch):
        self.ws_epoch = ws_epoch

    def _coerce_date(self, v) -> Optional[dt.date]:
        if v is None:
            return None
        if isinstance(v, dt.datetime):
            return v.date()
        if isinstance(v, dt.date):
            return v
        if isinstance(v, (int, float)):
            # Convert Excel serial to datetime using workbook epoch
            try:
                d = from_excel(v, self.ws_epoch)
            except Exception as exc:  # noqa: BLE001
                logger.warning("Failed converting Excel date %r: %s", v, exc)
                return None
            return d.date() if isinstance(d, dt.datetime) else d
        # Unsupported type
        logger.debug("Unsupported date cell type: %r", type(v))
        return None

    def _coerce_price(self, v) -> Optional[Decimal]:
        if v is None:
            return None
        if isinstance(v, Decimal):
            return v
        if isinstance(v, (int, float)):
            try:
                return Decimal(str(v))
            except Exception as exc:  # noqa: BLE001
                logger.warning("Failed converting price %r to Decimal: %s", v, exc)
                return None
        if isinstance(v, str):
            s = v.strip()
            if not s:
                return None
            # Defensive: handle locales with comma and spaces
            s = re.sub(r"[,\s]", "", s)
            try:
                return Decimal(s)
            except Exception as exc:  # noqa: BLE001
                logger.warning(
                    "Failed converting price string %r to Decimal: %s", v, exc
                )
                return None
        logger.debug("Unsupported price cell type: %r", type(v))
        return None

    def parse_rows(
        self, ws: Worksheet, currency_map: dict[int, str], start_row: int = 5
    ) -> list[ForexRecord]:
        records: list[ForexRecord] = []
        for row_idx in range(start_row, ws.max_row + 1):
            date_cell = ws.cell(row=row_idx, column=1)
            date_value = self._coerce_date(date_cell.value)
            if not date_value:
                # Stop on first empty date row; usually trailing empty rows.
                continue
            for col_idx, code in currency_map.items():
                price_cell = ws.cell(row=row_idx, column=col_idx)
                price = self._coerce_price(price_cell.value)
                if price is None:
                    continue
                records.append(ForexRecord(date=date_value, currency=code, price=price))
        logger.info("Parsed %d records from worksheet.", len(records))
        return records


class BdpForexParser:
    """High-level parser that coordinates header and row parsing."""

    def __init__(self) -> None:
        self._header = HeaderParser()

    def parse(self, xlsx_path: Path) -> list[ForexRecord]:
        wb = load_workbook(filename=str(xlsx_path), data_only=True)
        ws = wb.active
        currencies = self._header.build_currency_map(ws)
        rows = RowParser(wb.epoch).parse_rows(ws, currencies)
        return rows
