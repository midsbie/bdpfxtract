from __future__ import annotations

import argparse
import csv
import datetime as dt
import re
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.datetime import from_excel

from bdpfxtract.logging.logging import configure_logging


logger = configure_logging()


@dataclass(frozen=True)
class ForexRecord:
    date: dt.date
    currency: str
    price: Decimal


class HeaderParser:
    """Parses the worksheet header to map column indexes to currency codes.

    Expected layout (as observed in Banco de Portugal downloads):
    - Row 1: Series IDs (ignored)
    - Row 2: Series titles like "Dólar dos EUA (USD) - diário"
    - Row 3: "Métrica" / "Valor" (ignored)
    - Row 4: "Unidade de Medida" / "Divisa" (ignored)
    - Row 5+: Data rows: Col A is date, Col B.. are numeric values per currency.
    """

    CODE_RE = re.compile(r"\(([A-Z]{3})\)")

    def build_currency_map(self, ws: Worksheet) -> Dict[int, str]:
        mapping: Dict[int, str] = {}
        header_row = 2
        # Start from column 2 (B) because column 1 (A) is the date column.
        for col_idx in range(2, ws.max_column + 1):
            cell = ws.cell(row=header_row, column=col_idx)
            value = cell.value
            if not value:
                continue
            if not isinstance(value, str):
                continue
            m = self.CODE_RE.search(value)
            if not m:
                # No code found; skip this column to be safe.
                logger.debug("Skipping column %s without ISO code in header: %r", col_idx, value)
                continue
            code = m.group(1)
            mapping[col_idx] = code
        if not mapping:
            raise ValueError("No currency headers found in row 2.")
        logger.info("Detected %d currency columns.", len(mapping))
        return mapping


class RowParser:
    """Parses data rows into ForexRecord entities."""

    def __init__(self, ws_epoch):
        self.ws_epoch = ws_epoch

    def _coerce_date(self, v) -> Optional[dt.date]:
        if v is None:
            return None
        if isinstance(v, dt.datetime):
            return v.date()
        if isinstance(v, dt.date):
            return v
        if isinstance(v, (int, float)):
            # Convert Excel serial to datetime using workbook epoch
            try:
                d = from_excel(v, self.ws_epoch)
            except Exception as exc:  # noqa: BLE001
                logger.warning("Failed converting Excel date %r: %s", v, exc)
                return None
            return d.date() if isinstance(d, dt.datetime) else d
        # Unsupported type
        logger.debug("Unsupported date cell type: %r", type(v))
        return None

    def _coerce_price(self, v) -> Optional[Decimal]:
        if v is None:
            return None
        if isinstance(v, Decimal):
            return v
        if isinstance(v, (int, float)):
            try:
                return Decimal(str(v))
            except Exception as exc:  # noqa: BLE001
                logger.warning("Failed converting price %r to Decimal: %s", v, exc)
                return None
        if isinstance(v, str):
            s = v.strip()
            if not s:
                return None
            # Defensive: handle locales with comma and spaces
            s = re.sub(r"[,\s]", "", s)
            try:
                return Decimal(s)
            except Exception as exc:  # noqa: BLE001
                logger.warning("Failed converting price string %r to Decimal: %s", v, exc)
                return None
        logger.debug("Unsupported price cell type: %r", type(v))
        return None

    def parse_rows(self, ws: Worksheet, currency_map: Dict[int, str], start_row: int = 5) -> List[ForexRecord]:
        records: List[ForexRecord] = []
        for row_idx in range(start_row, ws.max_row + 1):
            date_cell = ws.cell(row=row_idx, column=1)
            date_value = self._coerce_date(date_cell.value)
            if not date_value:
                # Stop on first empty date row; usually trailing empty rows.
                continue
            for col_idx, code in currency_map.items():
                price_cell = ws.cell(row=row_idx, column=col_idx)
                price = self._coerce_price(price_cell.value)
                if price is None:
                    continue
                records.append(ForexRecord(date=date_value, currency=code, price=price))
        logger.info("Parsed %d records from worksheet.", len(records))
        return records


class CsvRepository:
    """Loads and persists ForexRecord items to CSV with merge semantics."""

    header = ("Date", "Currency", "Price")

    def load(self, path: Path) -> Dict[Tuple[str, str], Decimal]:
        data: Dict[Tuple[str, str], Decimal] = {}
        if not path.exists():
            return data
        with path.open("r", newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            missing_cols = [c for c in self.header if c not in reader.fieldnames or reader.fieldnames is None]
            if missing_cols:
                raise ValueError(f"CSV missing required columns: {missing_cols}")
            for row in reader:
                k = (row["Date"], row["Currency"])  # type: ignore[index]
                try:
                    data[k] = Decimal(row["Price"])  # type: ignore[index]
                except Exception as exc:  # noqa: BLE001
                    logger.warning("Skipping invalid CSV row %r: %s", row, exc)
        return data

    def save(self, path: Path, kv: Dict[Tuple[str, str], Decimal]) -> None:
        # Sort for stable output: by date then currency
        rows = sorted(kv.items(), key=lambda x: (x[0][0], x[0][1]))
        with path.open("w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=list(self.header))
            writer.writeheader()
            for (date_str, currency), price in rows:
                writer.writerow({
                    "Date": date_str,
                    "Currency": currency,
                    "Price": str(price),
                })

    @staticmethod
    def key_for(r: ForexRecord) -> Tuple[str, str]:
        return (r.date.isoformat(), r.currency)


class BdpForexParser:
    """High-level parser that coordinates header and row parsing."""

    def __init__(self) -> None:
        self._header = HeaderParser()

    def parse(self, xlsx_path: Path) -> List[ForexRecord]:
        wb = load_workbook(filename=str(xlsx_path), data_only=True)
        ws = wb.active
        currencies = self._header.build_currency_map(ws)
        rows = RowParser(wb.epoch).parse_rows(ws, currencies)
        return rows


class BdpForexToCsvApp:
    """Application entry point to parse an XLSX and merge/write a CSV.

    Usage:
      python -m bdpfxtract.cmd.bdp_forex_to_csv --input sample.xlsx --output out.csv
    """

    def __init__(self) -> None:
        self._parser = BdpForexParser()
        self._repo = CsvRepository()

    def run(self, xlsx_path: Path, out_csv: Path) -> Path:
        logger.info("Parsing workbook: %s", xlsx_path)
        records = self._parser.parse(xlsx_path)
        logger.info("Merging into CSV: %s", out_csv)

        existing = self._repo.load(out_csv)
        for r in records:
            existing[self._repo.key_for(r)] = r.price
        # Persist
        out_csv.parent.mkdir(parents=True, exist_ok=True)
        self._repo.save(out_csv, existing)
        logger.info("Wrote %d rows to %s", len(existing), out_csv)
        return out_csv


def build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        prog="bdp-forex-to-csv",
        description=(
            "Parse a Banco de Portugal FOREX XLSX and merge or create a CSV "
            "with columns: Date, Currency, Price."
        ),
    )
    p.add_argument("--input", "-i", required=True, help="Path to XLSX input file")
    p.add_argument("--output", "-o", required=True, help="Path to CSV output file")
    return p


def main(argv: Optional[Iterable[str]] = None) -> int:
    args = build_arg_parser().parse_args(argv)
    xlsx = Path(args.input)
    out_csv = Path(args.output)

    if not xlsx.exists():
        logger.error("Input XLSX not found: %s", xlsx)
        return 2

    try:
        BdpForexToCsvApp().run(xlsx, out_csv)
    except Exception as exc:  # noqa: BLE001
        logger.exception("Failed to process workbook: %s", exc)
        return 1
    return 0


if __name__ == "__main__":  # pragma: no cover
    raise SystemExit(main())

